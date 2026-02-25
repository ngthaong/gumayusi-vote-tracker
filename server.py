import os
import time
import json
import threading
from datetime import datetime, timezone, timedelta
from collections import deque

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import gspread
from google.oauth2.service_account import Credentials
from flask import Flask, jsonify, render_template, request as flask_request
from flask_cors import CORS

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

app = Flask(__name__, template_folder="templates", static_folder="static")
CORS(app)

# ── Configuration ────────────────────────────────────────────────────────────
POLL_ID = os.getenv("POLL_ID", "698ecd18c1bbe47a262c4f4b")
BSTAGE_EMAIL = os.getenv("BSTAGE_EMAIL", "ngthaongjenny@gmail.com")
BSTAGE_PASSWORD = os.getenv("BSTAGE_PASSWORD", "Guma123.")
FETCH_INTERVAL = int(os.getenv("FETCH_INTERVAL", "5" if os.getenv("VERCEL", "") else "3"))  # 5s on Vercel (includes API latency), 3s locally
WRITE_INTERVAL = int(os.getenv("WRITE_INTERVAL", "60"))  # Write to Excel/Google Sheet every 60 seconds
ENABLE_VERCEL_GSHEET_WRITE = os.getenv("ENABLE_VERCEL_GSHEET_WRITE", "true").lower() in ("1", "true", "yes")
SPACE_ID = "flnk-official"
VN_TZ = timezone(timedelta(hours=7))  # Vietnam timezone UTC+7
KST_TZ = timezone(timedelta(hours=9))  # Korean Standard Time UTC+9
IS_VERCEL = os.getenv("VERCEL", "") != ""  # True when running on Vercel (serverless)

# ── In-memory state ──────────────────────────────────────────────────────────
# Each snapshot: { "timestamp": ISO, "candidates": [...], "total": int }
vote_history = deque(maxlen=17280)  # ~24h of 5s snapshots

# Previous day winner (KST timezone)
_prev_day_data = {
    "date": None,       # "YYYY-MM-DD" in KST
    "winner": None,     # name of the leading candidate
    "winner_votes": 0,
    "runnerup": None,
    "runnerup_votes": 0,
    "diff": 0,
    "loaded": False,
}
_prev_day_kst_date = None  # Current KST date for day-change detection
_prev_day_loaded = False   # Whether we've tried loading from Google Sheet
_last_write_time = 0.0     # When we last wrote to Excel/Sheet (for throttling)
_last_write_snapshot = None  # {"doran": int, "guma": int} at last write, for diff calculation
current_data = {
    "poll_title": "",
    "poll_body": "",
    "poll_image": "",
    "candidates": [],
    "total_votes": 0,
    "last_updated": None,
    "error": None,
}
lock = threading.Lock()

# Auth state
access_token = None
refresh_token = None
token_expiry = 0


# ── BStage Authentication ────────────────────────────────────────────────────

def authenticate():
    """Full OAuth login flow: account.bstage.in → bstageplus.com token."""
    global access_token, refresh_token, token_expiry

    print(f"[{now()}] Authenticating...")

    try:
        s = requests.Session()
        s.headers.update({
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
        })

        # Step 1: Visit OAuth authorize page (sets cookies)
        s.get(
            "https://account.bstage.in/oauth/authorize"
            "?clientId=space_flnk-official&responseType=code"
            "&scope=openid%20profile%20email&state=state"
            "&redirectUri=https://bstageplus.com/account/login-callback/bstage",
            timeout=15,
        )

        # Step 2: Login at account.bstage.in
        r1 = s.post(
            "https://account.bstage.in/api/v1/bstage/auth/login",
            json={"email": BSTAGE_EMAIL, "password": BSTAGE_PASSWORD},
            headers={"Content-Type": "application/json", "Accept": "application/json"},
            timeout=15,
        )
        if r1.status_code != 200:
            raise Exception(f"Login failed: {r1.status_code} - {r1.text[:200]}")

        account_token = r1.json()["accessToken"]

        # Step 3: Get OAuth authorization code
        r2 = s.post(
            "https://account.bstage.in/api/v1/bstage/oauth/authorize",
            json={
                "clientId": "space_flnk-official",
                "responseType": "code",
                "scope": "openid profile email",
                "state": "state",
                "redirectUri": "https://bstageplus.com/account/login-callback/bstage",
            },
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {account_token}",
            },
            timeout=15,
        )
        if r2.status_code != 200:
            raise Exception(f"OAuth authorize failed: {r2.status_code} - {r2.text[:200]}")

        code = r2.json()["code"]

        # Step 4: Exchange code for bstageplus.com tokens
        r3 = s.post(
            "https://bstageplus.com/svc/account/api/v1/auth/token/social",
            json={
                "code": code,
                "state": "state",
                "socialType": "BSTAGE",
                "redirectUri": "https://bstageplus.com/account/login-callback/bstage",
            },
            headers={
                "Content-Type": "application/json",
                "Accept": "application/json",
                "x-bmf-sid": SPACE_ID,
            },
            timeout=15,
        )
        if r3.status_code != 200:
            raise Exception(f"Token exchange failed: {r3.status_code} - {r3.text[:200]}")

        token_data = r3.json()
        access_token = token_data["accessToken"]
        refresh_token = token_data.get("refreshToken", "")
        # Token expires in ~30 minutes, refresh at 25 min
        token_expiry = time.time() + 25 * 60

        print(f"[{now()}] ✓ Authentication successful!")
        return True

    except Exception as e:
        print(f"[{now()}] ✗ Authentication failed: {e}")
        with lock:
            current_data["error"] = f"Authentication failed: {e}"
        return False


def ensure_auth():
    """Ensure we have a valid access token, re-authenticate if needed."""
    global access_token
    if not access_token or time.time() > token_expiry:
        return authenticate()
    return True


def try_refresh_token():
    """Try to refresh the access token without full re-login."""
    global access_token, refresh_token, token_expiry

    if not refresh_token:
        return authenticate()

    try:
        r = requests.post(
            "https://bstageplus.com/svc/account/api/v1/auth/token/refresh",
            json={"accessToken": access_token, "refreshToken": refresh_token},
            headers={
                "Content-Type": "application/json",
                "Accept": "application/json",
                "x-bmf-sid": SPACE_ID,
            },
            timeout=15,
        )
        if r.status_code == 200:
            data = r.json()
            access_token = data.get("accessToken", access_token)
            refresh_token = data.get("refreshToken", refresh_token)
            token_expiry = time.time() + 25 * 60
            print(f"[{now()}] ✓ Token refreshed")
            return True
    except Exception:
        pass

    # Fallback to full re-authentication
    return authenticate()


# ── Data Fetching ────────────────────────────────────────────────────────────

def fetch_poll_metadata():
    """Fetch poll title, description, and candidate list (no auth needed)."""
    try:
        r = requests.get(
            f"https://bstageplus.com/svc/survey/api/v1/polls/{POLL_ID}",
            headers={
                "User-Agent": "Mozilla/5.0",
                "Accept": "application/json",
                "x-bmf-sid": SPACE_ID,
            },
            timeout=15,
        )
        if r.status_code == 200:
            return r.json()
    except Exception as e:
        print(f"[{now()}] Metadata fetch error: {e}")
    return None


def fetch_poll_results():
    """Fetch poll voting results (requires auth)."""
    if not ensure_auth():
        return None

    try:
        r = requests.get(
            f"https://bstageplus.com/svc/survey/api/v1/polls/{POLL_ID}/results",
            headers={
                "User-Agent": "Mozilla/5.0",
                "Accept": "application/json",
                "x-bmf-sid": SPACE_ID,
                "Authorization": f"Bearer {access_token}",
            },
            timeout=15,
        )
        if r.status_code == 200:
            return r.json()
        elif r.status_code == 401:
            print(f"[{now()}] Token expired, refreshing...")
            if try_refresh_token():
                return fetch_poll_results()  # Retry once
        else:
            print(f"[{now()}] Results fetch error: {r.status_code} - {r.text[:200]}")
    except Exception as e:
        print(f"[{now()}] Results fetch error: {e}")
    return None


def fetch_poll_data():
    """Main fetch: get metadata + results and update state."""
    global current_data, _last_write_time, _last_write_snapshot

    print(f"[{now()}] Fetching poll data...")

    metadata = fetch_poll_metadata()
    results = fetch_poll_results()

    if results:
        candidates = []
        option_results = results.get("questionOptionResults", [])

        for opt in option_results:
            candidates.append({
                "id": opt.get("questionOptionId", ""),
                "name": opt.get("content", "Unknown"),
                "image": opt.get("images", [""])[0] if opt.get("images") else "",
                "votes": opt.get("numberOfSelectors", 0),
            })

        total = sum(c["votes"] for c in candidates)
        timestamp = datetime.now(VN_TZ).isoformat()

        snapshot = {
            "timestamp": timestamp,
            "candidates": candidates,
            "total": total,
        }

        with lock:
            vote_history.append(snapshot)
            current_data.update({
                "poll_title": metadata.get("title", "") if metadata else current_data.get("poll_title", ""),
                "poll_body": metadata.get("body", "") if metadata else current_data.get("poll_body", ""),
                "poll_image": metadata.get("mainImage", "") if metadata else current_data.get("poll_image", ""),
                "candidates": candidates,
                "total_votes": total,
                "last_updated": timestamp,
                "error": None,
            })

        print(f"[{now()}] ✓ {total:,} total votes across {len(candidates)} candidates")

        # Write results to xlsx and Google Sheet every WRITE_INTERVAL seconds (skip xlsx on Vercel — no filesystem)
        now_ts = time.time()
        should_write = (now_ts - _last_write_time >= WRITE_INTERVAL) if not IS_VERCEL else (
            _gsheet_should_write() if ENABLE_VERCEL_GSHEET_WRITE else False
        )
        if should_write:
            if not IS_VERCEL:
                write_result_file()
            write_google_sheet()
            _last_write_time = now_ts
            with lock:
                cands = current_data.get("candidates", [])
            tracked = {c["name"].strip(): c["votes"] for c in cands if c["name"].strip() in TRACK_NAMES}
            if len(tracked) >= 2:
                _last_write_snapshot = {"doran": tracked.get("T1 Doran", 0), "guma": tracked.get("Hanwha Life Esports Gumayusi", 0)}
    else:
        with lock:
            if not current_data.get("error"):
                current_data["error"] = "Failed to fetch poll results. Retrying..."
        print(f"[{now()}] ✗ Failed to fetch poll data")


# ── Result File Writer ────────────────────────────────────────────────────────

TRACK_NAMES = ["T1 Doran", "Hanwha Life Esports Gumayusi"]
RESULT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "result.xlsx")

def _init_xlsx():
    """Create a new xlsx file with headers and styling."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vote Tracker"

    # Header styles
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    headers = ["Time", "Doran Votes", "Doran (+)", "Gumayusi Votes", "Gumayusi (+)", "Gap (D−G)"]
    col_widths = [22, 15, 12, 17, 14, 14]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # Freeze the header row
    ws.freeze_panes = "A2"

    wb.save(RESULT_PATH)
    return wb


def write_result_file():
    """Append a row to result.xlsx with Doran & Gumayusi votes."""
    try:
        with lock:
            candidates = current_data.get("candidates", [])
            history = list(vote_history)

        if not candidates:
            return

        # Find the two candidates (strip whitespace for matching)
        tracked = {}
        for c in candidates:
            if c["name"].strip() in TRACK_NAMES:
                tracked[c["name"].strip()] = c["votes"]

        if len(tracked) < 2:
            return

        doran = tracked.get("T1 Doran", 0)
        guma = tracked.get("Hanwha Life Esports Gumayusi", 0)
        gap = doran - guma
        # Use fetch timestamp (not write time) for consistent ordering; fallback to now if missing
        ts_iso = current_data.get("last_updated")
        try:
            timestamp = datetime.fromisoformat(ts_iso).strftime("%Y-%m-%d %H:%M:%S") if ts_iso else datetime.now(VN_TZ).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            timestamp = datetime.now(VN_TZ).strftime("%Y-%m-%d %H:%M:%S")

        # Calculate per-candidate change (since last write for 1-min diff)
        doran_diff = 0
        guma_diff = 0
        if _last_write_snapshot:
            doran_diff = doran - _last_write_snapshot.get("doran", doran)
            guma_diff = guma - _last_write_snapshot.get("guma", guma)
        elif len(history) >= 2:
            prev_map = {c["name"].strip(): c["votes"] for c in history[-2]["candidates"]}
            doran_diff = doran - prev_map.get("T1 Doran", doran)
            guma_diff = guma - prev_map.get("Hanwha Life Esports Gumayusi", guma)

        # Load or create workbook
        if os.path.exists(RESULT_PATH):
            wb = load_workbook(RESULT_PATH)
        else:
            wb = _init_xlsx()

        ws = wb.active
        next_row = ws.max_row + 1

        # Cell styles
        data_font = Font(name="Arial", size=10)
        num_font = Font(name="Arial", size=10, bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        green_font = Font(name="Arial", size=10, bold=True, color="059669")
        red_font = Font(name="Arial", size=10, bold=True, color="DC2626")
        thin_border = Border(
            left=Side(style="thin", color="EEEEEE"),
            right=Side(style="thin", color="EEEEEE"),
            top=Side(style="thin", color="EEEEEE"),
            bottom=Side(style="thin", color="EEEEEE"),
        )

        # Alternate row color
        row_fill = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid") if next_row % 2 == 0 else PatternFill(fill_type=None)

        # Write row
        row_data = [timestamp, doran, doran_diff, guma, guma_diff, gap]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            cell.border = thin_border
            cell.fill = row_fill

            if col_idx == 1:
                cell.font = data_font
                cell.alignment = center_align
            elif col_idx in (2, 4):  # vote counts
                cell.font = num_font
                cell.alignment = right_align
                cell.number_format = "#,##0"
            elif col_idx in (3, 5):  # diffs
                cell.alignment = center_align
                cell.number_format = "+#,##0;-#,##0;0"
                cell.font = green_font if val > 0 else (red_font if val < 0 else data_font)
            elif col_idx == 6:  # gap
                cell.alignment = center_align
                cell.number_format = "+#,##0;-#,##0;0"
                cell.font = Font(name="Arial", size=10, bold=True, color="EA580C")

        wb.save(RESULT_PATH)

        gap_str = f"+{gap:,}" if gap > 0 else f"{gap:,}" if gap < 0 else "TIE"
        print(f"[{now()}] 📊 result.xlsx row {next_row}: Doran {doran:,} vs Guma {guma:,} (gap: {gap_str})")

    except Exception as e:
        print(f"[{now()}] Failed to write result.xlsx: {e}")


# ── Google Sheets Writer ──────────────────────────────────────────────────────

GOOGLE_SHEET_ID = "1BiS63WJqQSICd1oBkSBAHpEF0syrXKq0Zcw_zTMicGg"

# Google creds: prefer env var GOOGLE_CREDS_JSON, fallback to file
_google_creds_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "braided-box-488323-g1-46311e11eceb.json")
_google_creds_json_env = os.getenv("GOOGLE_CREDS_JSON", "")
if _google_creds_json_env:
    # Write env var to a temp file so google-auth can read it
    import tempfile
    _tmp_creds = tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False)
    _tmp_creds.write(_google_creds_json_env)
    _tmp_creds.close()
    GOOGLE_CREDS_FILE = _tmp_creds.name
elif os.path.exists(_google_creds_file_path):
    GOOGLE_CREDS_FILE = _google_creds_file_path
else:
    GOOGLE_CREDS_FILE = None
_gsheet = None
_gsheet_spreadsheet = None
_gsheet_throttle = None  # Dedicated sheet for last-write timestamp (avoids G1 conflicts)
_gsheet_initialized = False
_gsheet_row_counter = 0  # Track rows locally to avoid extra API reads
_gsheet_queue = deque(maxlen=500)  # Queue for async writes
_gsheet_lock = threading.Lock()

def _init_gsheet():
    """Initialize Google Sheets connection."""
    global _gsheet, _gsheet_spreadsheet, _gsheet_throttle, _gsheet_initialized, _gsheet_row_counter
    if not GOOGLE_CREDS_FILE:
        print(f"[{now()}] ✗ Google creds not found, skipping Google Sheets")
        return False
    try:
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        creds = Credentials.from_service_account_file(GOOGLE_CREDS_FILE, scopes=scopes)
        gc = gspread.authorize(creds)
        _gsheet_spreadsheet = gc.open_by_key(GOOGLE_SHEET_ID)
        _gsheet = _gsheet_spreadsheet.sheet1

        # Get or create "_throttle" sheet for last-write timestamp (Vercel throttling)
        try:
            _gsheet_throttle = _gsheet_spreadsheet.worksheet("_throttle")
        except gspread.WorksheetNotFound:
            _gsheet_throttle = _gsheet_spreadsheet.add_worksheet(title="_throttle", rows=1, cols=1)
            _gsheet_throttle.update_acell("A1", "0")

        # Write header if sheet is empty
        if _gsheet.row_count == 0 or not _gsheet.cell(1, 1).value:
            _gsheet.update(values=[["Time", "Doran Votes", "Doran (+)", "Gumayusi Votes", "Gumayusi (+)", "Gap (D−G)"]], range_name="A1:F1")
            _gsheet.format("A1:F1", {
                "backgroundColor": {"red": 0.976, "green": 0.451, "blue": 0.086},
                "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}, "fontSize": 11},
                "horizontalAlignment": "CENTER",
            })
            _gsheet.freeze(rows=1)

        # Count existing rows once
        all_vals = _gsheet.col_values(1)
        _gsheet_row_counter = len(all_vals)

        _gsheet_initialized = True
        print(f"[{now()}] ✓ Google Sheet connected! ({_gsheet_row_counter} rows)")
        return True
    except Exception as e:
        print(f"[{now()}] ✗ Google Sheet init failed: {e}")
        return False


def _gsheet_batch_format(sheet_id, row_idx):
    """Build a batch_update request to format a single row (1 API call instead of 6)."""
    return {
        "requests": [
            # Center align entire row
            {
                "repeatCell": {
                    "range": {"sheetId": sheet_id, "startRowIndex": row_idx - 1, "endRowIndex": row_idx, "startColumnIndex": 0, "endColumnIndex": 6},
                    "cell": {"userEnteredFormat": {"horizontalAlignment": "CENTER", "verticalAlignment": "MIDDLE"}},
                    "fields": "userEnteredFormat(horizontalAlignment,verticalAlignment)",
                }
            },
            # Number format for B (Doran votes), D (Guma votes), F (Gap)
            {
                "repeatCell": {
                    "range": {"sheetId": sheet_id, "startRowIndex": row_idx - 1, "endRowIndex": row_idx, "startColumnIndex": 1, "endColumnIndex": 2},
                    "cell": {"userEnteredFormat": {"numberFormat": {"type": "NUMBER", "pattern": "#,##0"}}},
                    "fields": "userEnteredFormat.numberFormat",
                }
            },
            {
                "repeatCell": {
                    "range": {"sheetId": sheet_id, "startRowIndex": row_idx - 1, "endRowIndex": row_idx, "startColumnIndex": 3, "endColumnIndex": 4},
                    "cell": {"userEnteredFormat": {"numberFormat": {"type": "NUMBER", "pattern": "#,##0"}}},
                    "fields": "userEnteredFormat.numberFormat",
                }
            },
            {
                "repeatCell": {
                    "range": {"sheetId": sheet_id, "startRowIndex": row_idx - 1, "endRowIndex": row_idx, "startColumnIndex": 5, "endColumnIndex": 6},
                    "cell": {"userEnteredFormat": {"numberFormat": {"type": "NUMBER", "pattern": "#,##0"}}},
                    "fields": "userEnteredFormat.numberFormat",
                }
            },
            # Diff format for C (Doran +), E (Guma +)
            {
                "repeatCell": {
                    "range": {"sheetId": sheet_id, "startRowIndex": row_idx - 1, "endRowIndex": row_idx, "startColumnIndex": 2, "endColumnIndex": 3},
                    "cell": {"userEnteredFormat": {"numberFormat": {"type": "NUMBER", "pattern": "+#,##0;-#,##0;0"}}},
                    "fields": "userEnteredFormat.numberFormat",
                }
            },
            {
                "repeatCell": {
                    "range": {"sheetId": sheet_id, "startRowIndex": row_idx - 1, "endRowIndex": row_idx, "startColumnIndex": 4, "endColumnIndex": 5},
                    "cell": {"userEnteredFormat": {"numberFormat": {"type": "NUMBER", "pattern": "+#,##0;-#,##0;0"}}},
                    "fields": "userEnteredFormat.numberFormat",
                }
            },
        ]
    }


def _gsheet_writer_thread():
    """Dedicated thread for writing to Google Sheet. Prevents blocking the data fetcher."""
    global _gsheet, _gsheet_spreadsheet, _gsheet_initialized, _gsheet_row_counter
    while True:
        try:
            # Wait for items in the queue
            if not _gsheet_queue:
                time.sleep(0.5)
                continue

            if not _gsheet_initialized:
                if not _init_gsheet():
                    time.sleep(10)  # Wait before retrying
                    continue

            with _gsheet_lock:
                if not _gsheet_queue:
                    continue
                row = _gsheet_queue.popleft()

            # Skip if last row same minute (dedupe)
            if _gsheet_last_row_same_minute(row[0] if row else ""):
                continue

            # Append the row (1 API call)
            _gsheet.append_row(row, value_input_option="RAW")
            _gsheet_row_counter += 1

            # Format the row using batch_update (1 API call instead of 6)
            sheet_id = _gsheet._properties.get("sheetId", 0)
            body = _gsheet_batch_format(sheet_id, _gsheet_row_counter)
            _gsheet_spreadsheet.batch_update(body)

            # Update _throttle sheet for Vercel throttling (shared across local + Vercel)
            if _gsheet_throttle:
                try:
                    _gsheet_throttle.update_acell("A1", str(int(time.time())))
                except Exception:
                    pass

        except Exception as e:
            err_str = str(e)
            if "429" in err_str or "Quota" in err_str:
                # Rate limited — back off for 60 seconds, don't drop the row
                print(f"[{now()}] Google Sheet rate limited, backing off 60s...")
                time.sleep(60)
            else:
                print(f"[{now()}] Google Sheet write error: {e}")
                _gsheet_initialized = False
                time.sleep(5)


def _gsheet_should_write():
    """On Vercel: check _throttle sheet A1 for last write timestamp. Returns True if we should write."""
    if not IS_VERCEL:
        return True  # Local uses _last_write_time in fetch_poll_data
    try:
        if not _gsheet_initialized and not _init_gsheet():
            return True  # Can't check, allow write
        if not _gsheet_throttle:
            return True
        val = (_gsheet_throttle.acell("A1").value or "").strip().replace(",", "")
        if not val or not val.replace(".", "").isdigit():
            return True  # No previous write or invalid
        last_ts = float(val)
        return (time.time() - last_ts) >= WRITE_INTERVAL
    except Exception:
        return True  # On error, allow write


def _gsheet_last_row_same_minute(new_ts_str):
    """Check if last data row has same minute as new_ts_str. Skip write if so (dedupe)."""
    try:
        if not _gsheet_initialized or not _gsheet:
            return False
        last_row = _gsheet.row_count
        if last_row < 2:
            return False
        last_ts = (_gsheet.cell(last_row, 1).value or "").strip()
        if not last_ts or not new_ts_str:
            return False
        # Compare minute: "2026-02-25 16:29:30" -> "2026-02-25 16:29"
        def minute_key(s):
            return s[:16] if len(s) >= 16 else s
        return minute_key(last_ts) == minute_key(new_ts_str)
    except Exception:
        return False


def write_google_sheet():
    """Queue a row for async Google Sheet writing (non-blocking)."""
    try:
        with lock:
            candidates = current_data.get("candidates", [])
            history = list(vote_history)

        if not candidates:
            return

        # Find the two candidates
        tracked = {}
        for c in candidates:
            if c["name"].strip() in TRACK_NAMES:
                tracked[c["name"].strip()] = c["votes"]

        if len(tracked) < 2:
            return

        doran = tracked.get("T1 Doran", 0)
        guma = tracked.get("Hanwha Life Esports Gumayusi", 0)
        gap = doran - guma
        # Use fetch timestamp (not write time) for consistent ordering; fallback to now if missing
        ts_iso = current_data.get("last_updated")
        try:
            timestamp = datetime.fromisoformat(ts_iso).strftime("%Y-%m-%d %H:%M:%S") if ts_iso else datetime.now(VN_TZ).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            timestamp = datetime.now(VN_TZ).strftime("%Y-%m-%d %H:%M:%S")

        # Calculate change (since last write for 1-min diff)
        doran_diff = 0
        guma_diff = 0
        if _last_write_snapshot:
            doran_diff = doran - _last_write_snapshot.get("doran", doran)
            guma_diff = guma - _last_write_snapshot.get("guma", guma)
        elif len(history) >= 2:
            prev_map = {c["name"].strip(): c["votes"] for c in history[-2]["candidates"]}
            doran_diff = doran - prev_map.get("T1 Doran", doran)
            guma_diff = guma - prev_map.get("Hanwha Life Esports Gumayusi", guma)

        row = [timestamp, doran, doran_diff, guma, guma_diff, gap]

        # Skip if last row is same minute (prevents 2–3 rows/min from concurrent requests)
        if _gsheet_last_row_same_minute(timestamp):
            return

        if IS_VERCEL:
            # On Vercel: write directly (no background thread)
            if not _gsheet_initialized:
                if not _init_gsheet():
                    return
            try:
                _gsheet.append_row(row, value_input_option="RAW")
                _gsheet_row_counter += 1
                sheet_id = _gsheet._properties.get("sheetId", 0)
                body = _gsheet_batch_format(sheet_id, _gsheet_row_counter)
                _gsheet_spreadsheet.batch_update(body)
                # Store last write time in _throttle sheet (persists across cold starts)
                if _gsheet_throttle:
                    try:
                        _gsheet_throttle.update_acell("A1", str(int(time.time())))
                    except Exception:
                        pass
            except Exception as ex:
                print(f"[{now()}] Vercel GSheet write error: {ex}")
        else:
            # Queue the row (non-blocking) for background thread
            with _gsheet_lock:
                _gsheet_queue.append(row)

    except Exception as e:
        print(f"[{now()}] Google Sheet queue error: {e}")


# ── Previous Day Winner Tracking (KST) ────────────────────────────────────────

def _check_day_change():
    """Detect KST midnight crossing and save yesterday's final snapshot."""
    global _prev_day_kst_date, _prev_day_data

    now_kst = datetime.now(KST_TZ)
    today_kst = now_kst.strftime("%Y-%m-%d")

    if _prev_day_kst_date is None:
        _prev_day_kst_date = today_kst
        return

    if today_kst != _prev_day_kst_date:
        # Day changed! Save the current vote counts as yesterday's final result
        with lock:
            candidates = current_data.get("candidates", [])

        if candidates:
            sorted_c = sorted(candidates, key=lambda c: c["votes"], reverse=True)
            winner = sorted_c[0]
            runnerup = sorted_c[1] if len(sorted_c) > 1 else {"name": "N/A", "votes": 0}
            _prev_day_data = {
                "date": _prev_day_kst_date,
                "winner": winner["name"],
                "winner_votes": winner["votes"],
                "runnerup": runnerup["name"],
                "runnerup_votes": runnerup["votes"],
                "diff": winner["votes"] - runnerup["votes"],
                "loaded": True,
            }
            print(f"[{now()}] 📅 Day changed → {today_kst} KST. Yesterday's winner: {winner['name']} (by {winner['votes'] - runnerup['votes']:,})")

        _prev_day_kst_date = today_kst


def _load_prev_day_from_gsheet():
    """Load previous day's final data from Google Sheet on startup/cold-start."""
    global _prev_day_data, _prev_day_loaded, _prev_day_kst_date

    if _prev_day_loaded:
        return
    _prev_day_loaded = True

    try:
        if not _gsheet_initialized:
            if not _init_gsheet():
                return

        now_kst = datetime.now(KST_TZ)
        today_kst = now_kst.strftime("%Y-%m-%d")
        _prev_day_kst_date = today_kst
        yesterday_kst = (now_kst - timedelta(days=1)).strftime("%Y-%m-%d")

        # Read all timestamps from column A
        all_ts = _gsheet.col_values(1)
        total = len(all_ts)
        if total <= 1:
            print(f"[{now()}] ⚠ Google Sheet too small to find yesterday's data")
            return

        # Search backwards for the last entry from yesterday (KST)
        # Timestamps are in VN (UTC+7), KST = VN + 2 hours
        target_row = None
        for i in range(total - 1, 0, -1):
            try:
                ts_vn = datetime.strptime(all_ts[i], "%Y-%m-%d %H:%M:%S").replace(tzinfo=VN_TZ)
                ts_kst = ts_vn.astimezone(KST_TZ)
                row_date = ts_kst.strftime("%Y-%m-%d")
                if row_date == yesterday_kst:
                    target_row = i + 1  # 1-indexed for gsheet
                    break
                elif row_date < yesterday_kst:
                    break
            except (ValueError, IndexError):
                continue

        if target_row is None:
            print(f"[{now()}] ⚠ No data found for yesterday ({yesterday_kst} KST)")
            return

        # Read just that row: [timestamp, doran, doran_diff, guma, guma_diff, gap]
        row_data = _gsheet.row_values(target_row)
        if len(row_data) < 4:
            return

        def parse_int(val):
            """Parse int from Google Sheet value (handles 2.596.698 or 2,596,698 or plain)."""
            s = str(val).strip()
            # Remove thousand separators (both . and ,)
            # If the string has dots as thousand seps (e.g. 2.596.698), remove them
            if s.count('.') > 1:
                s = s.replace('.', '')
            elif '.' in s and ',' in s:
                s = s.replace(',', '')  # e.g. 2,596.698 (unlikely but handle it)
            else:
                s = s.replace(',', '').replace('.', '')
            return int(s)

        doran = parse_int(row_data[1])
        guma = parse_int(row_data[3])

        if doran >= guma:
            _prev_day_data = {
                "date": yesterday_kst,
                "winner": "T1 Doran",
                "winner_votes": doran,
                "runnerup": "HLE Gumayusi",
                "runnerup_votes": guma,
                "diff": doran - guma,
                "loaded": True,
            }
        else:
            _prev_day_data = {
                "date": yesterday_kst,
                "winner": "HLE Gumayusi",
                "winner_votes": guma,
                "runnerup": "T1 Doran",
                "runnerup_votes": doran,
                "diff": guma - doran,
                "loaded": True,
            }
        print(f"[{now()}] 📅 Loaded yesterday's winner from Google Sheet: {_prev_day_data['winner']} (by {_prev_day_data['diff']:,}) [{yesterday_kst} KST]")

    except Exception as e:
        print(f"[{now()}] Error loading yesterday's data: {e}")


# ── Background Worker ────────────────────────────────────────────────────────

def background_fetcher():
    """Periodically fetch poll data."""
    while True:
        try:
            _check_day_change()
            fetch_poll_data()
        except Exception as e:
            print(f"[{now()}] Background fetcher error: {e}")
        time.sleep(FETCH_INTERVAL)


# ── Utility ──────────────────────────────────────────────────────────────────

def now():
    return datetime.now(VN_TZ).strftime("%H:%M:%S")


# ── API Routes ───────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/current")
def api_current():
    """Current poll data with vote velocity."""
    # On Vercel (serverless): fetch data on-demand if not fresh
    if IS_VERCEL:
        # Load yesterday's data from Google Sheet on cold start
        if not _prev_day_loaded:
            try:
                _load_prev_day_from_gsheet()
            except Exception as e:
                print(f"[{now()}] Prev day load error: {e}")

        with lock:
            last = current_data.get("last_updated")
        need_fetch = True
        if last:
            try:
                age = (datetime.now(VN_TZ) - datetime.fromisoformat(last)).total_seconds()
                need_fetch = age > 5  # only re-fetch if data is older than 5s
            except Exception:
                pass
        if need_fetch:
            try:
                _check_day_change()
                fetch_poll_data()
            except Exception as e:
                print(f"[{now()}] On-demand fetch error: {e}")

    with lock:
        data = dict(current_data)

        # Find the 1st place votes for gap calculation
        first_place_votes = 0
        if data.get("candidates"):
            first_place_votes = max(c["votes"] for c in data["candidates"])

        if len(vote_history) >= 2:
            latest = vote_history[-1]
            previous = vote_history[-2]

            t_latest = datetime.fromisoformat(latest["timestamp"])
            t_prev = datetime.fromisoformat(previous["timestamp"])
            minutes = max((t_latest - t_prev).total_seconds() / 60.0, 0.0167)

            prev_map = {c["name"]: c["votes"] for c in previous["candidates"]}

            enriched = []
            for c in data["candidates"]:
                prev_votes = prev_map.get(c["name"], c["votes"])
                diff = c["votes"] - prev_votes
                velocity = round(diff / minutes, 1)
                gap = first_place_votes - c["votes"]
                enriched.append({**c, "velocity": velocity, "diff": diff, "gap_from_first": gap})

            data["candidates"] = enriched
            data["total_diff"] = latest["total"] - previous["total"]
            data["total_velocity"] = round(
                (latest["total"] - previous["total"]) / minutes, 1
            )
        else:
            for c in data.get("candidates", []):
                c["velocity"] = 0
                c["diff"] = 0
                c["gap_from_first"] = first_place_votes - c["votes"]
            data["total_diff"] = 0
            data["total_velocity"] = 0

        data["fetch_interval"] = FETCH_INTERVAL
        data["history_length"] = len(vote_history)
        data["previous_day"] = _prev_day_data if _prev_day_data.get("loaded") else None

    return jsonify(data)


@app.route("/api/history")
def api_history():
    """Vote history for charting."""
    with lock:
        return jsonify(list(vote_history))


@app.route("/api/refresh", methods=["POST"])
def api_refresh():
    """Trigger immediate data refresh."""
    threading.Thread(target=fetch_poll_data, daemon=True).start()
    return jsonify({"status": "ok", "message": "Refresh triggered"})


# ── Main ─────────────────────────────────────────────────────────────────────

# ── Start background threads (works with both direct run and gunicorn) ────────
_threads_started = False

def start_background_threads():
    global _threads_started
    if _threads_started:
        return
    _threads_started = True

    print("=" * 60)
    print("  🗳️  BStage Plus Vote Tracker")
    print("=" * 60)
    print(f"  Poll ID:        {POLL_ID}")
    print(f"  Fetch Interval: {FETCH_INTERVAL}s")
    print(f"  Account:        {BSTAGE_EMAIL}")
    print("=" * 60)

    # Load yesterday's winner from Google Sheet on startup
    try:
        _load_prev_day_from_gsheet()
    except Exception as e:
        print(f"  ⚠ Could not load yesterday's data: {e}")

    # Start background fetcher
    fetcher = threading.Thread(target=background_fetcher, daemon=True)
    fetcher.start()

    # Start Google Sheet writer thread (separate from fetcher to avoid blocking)
    gs_writer = threading.Thread(target=_gsheet_writer_thread, daemon=True)
    gs_writer.start()


# Auto-start threads when module is loaded (for gunicorn) — skip on Vercel (serverless)
if not IS_VERCEL:
    start_background_threads()
else:
    print("Running on Vercel (serverless mode) — no background threads")


if __name__ == "__main__":
    port = int(os.getenv("PORT", "5050"))
    print(f"  Port:           {port}")
    app.run(host="0.0.0.0", port=port, debug=False)
